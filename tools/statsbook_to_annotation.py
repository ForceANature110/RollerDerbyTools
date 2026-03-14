import json
import argparse
from pathlib import Path

from openpyxl import load_workbook


# -----------------------------
# Helpers
# -----------------------------

def as_text(value):
    if value is None:
        return ""
    return str(value).strip()


def as_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except Exception:
        try:
            return int(float(value))
        except Exception:
            return None


def cell_text(ws, coord):
    return as_text(ws[coord].value)


def cell_int(ws, coord):
    return as_int(ws[coord].value)


def nonempty_list(values):
    return [v for v in values if as_text(v) != ""]


def output_json_path(xlsx_path: Path, explicit_output: Path | None) -> Path:
    if explicit_output is not None:
        return explicit_output
    return xlsx_path.with_name(f"{xlsx_path.stem}.jams.json")


# -----------------------------
# Sheet mappings
# -----------------------------

SCORE_LAYOUT = {
    1: {"start_row": 4, "end_row": 41},
    2: {"start_row": 46, "end_row": 83},
}

HOME_SCORE_COLUMNS = {
    "jam_number": "A",
    "jammer": "B",
    "lead": "D",
    "trip_cols": ["H", "I", "J", "K", "L", "M", "N", "O", "P"],
}

AWAY_SCORE_COLUMNS = {
    "jam_number": "T",
    "jammer": "U",
    "lead": "W",
    "trip_cols": ["AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI"],
}

LINEUPS_LAYOUT = {
    1: {"start_row": 4, "end_row": 41},
    2: {"start_row": 46, "end_row": 83},
}

HOME_LINEUP_COLUMNS = ["G", "K", "O", "S"]
AWAY_LINEUP_COLUMNS = ["AG", "AK", "AO", "AS"]

PENALTY_COLUMN_GROUPS = {
    1: {
        "home": ["B", "C", "D", "E", "F", "G", "H", "I", "J"],
        "away": ["Q", "R", "S", "T", "U", "V", "W", "X", "Y"],
    },
    2: {
        "home": ["AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL"],
        "away": ["AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ", "BA"],
    },
}

PENALTY_ROSTER_COLUMNS = {
    1: {"home": "A", "away": "P"},
    2: {"home": "AC", "away": "AR"},
}

PENALTY_ROSTER_START_ROW = 4
PENALTY_ROSTER_END_ROW = 43
PENALTY_ROW_STEP = 2


# -----------------------------
# Read score sheet
# -----------------------------

def read_passes(ws, row, trip_cols):
    passes = []
    for col in trip_cols:
        value = ws[f"{col}{row}"].value
        if value in (None, ""):
            continue
        parsed = as_int(value)
        if parsed is not None:
            passes.append(parsed)
    return passes


def read_lead(home_lead, away_lead):
    home_marked = as_text(home_lead).upper() == "X"
    away_marked = as_text(away_lead).upper() == "X"

    if home_marked and not away_marked:
        return "home"
    if away_marked and not home_marked:
        return "away"
    if not home_marked and not away_marked:
        return "none"
    return "unknown"


def extract_jams_from_score(score_ws):
    jams = []

    for period_number, layout in SCORE_LAYOUT.items():
        for row in range(layout["start_row"], layout["end_row"] + 1):
            home_jam_number = cell_int(score_ws, f"{HOME_SCORE_COLUMNS['jam_number']}{row}")
            away_jam_number = cell_int(score_ws, f"{AWAY_SCORE_COLUMNS['jam_number']}{row}")

            jam_number = home_jam_number if home_jam_number is not None else away_jam_number
            if jam_number is None:
                continue

            jam = {
                "period_number": period_number,
                "jam_number": jam_number,
                "start_time": None,
                "end_time": None,
                "home_jammer": cell_text(score_ws, f"{HOME_SCORE_COLUMNS['jammer']}{row}"),
                "away_jammer": cell_text(score_ws, f"{AWAY_SCORE_COLUMNS['jammer']}{row}"),
                "home_lineup": [],
                "away_lineup": [],
                "home_score_start": None,
                "away_score_start": None,
                "home_passes": read_passes(score_ws, row, HOME_SCORE_COLUMNS["trip_cols"]),
                "away_passes": read_passes(score_ws, row, AWAY_SCORE_COLUMNS["trip_cols"]),
                "home_score_end": None,
                "away_score_end": None,
                "home_star_pass": False,
                "away_star_pass": False,
                "lead_jammer": read_lead(
                    score_ws[f"{HOME_SCORE_COLUMNS['lead']}{row}"].value,
                    score_ws[f"{AWAY_SCORE_COLUMNS['lead']}{row}"].value,
                ),
                "penalties": [],
                "notes": "",
            }
            jams.append(jam)

    jams.sort(key=lambda j: (j["period_number"], j["jam_number"]))
    return jams


# -----------------------------
# Read lineups sheet
# -----------------------------

def index_jams(jams):
    return {(j["period_number"], j["jam_number"]): j for j in jams}


def parse_jam_marker(value):
    text = as_text(value).upper()
    if text == "":
        return None
    if text == "SP":
        return "SP"
    return as_int(value)


def read_lineup_cells(ws, row, columns):
    return nonempty_list([ws[f"{col}{row}"].value for col in columns])


def extract_lineups(lineups_ws, jams_by_key):
    for period_number, layout in LINEUPS_LAYOUT.items():
        current_jam_number = 0

        for row in range(layout["start_row"], layout["end_row"] + 1):
            home_marker = parse_jam_marker(lineups_ws[f"A{row}"].value)
            away_marker = parse_jam_marker(lineups_ws[f"AA{row}"].value)

            home_lineup = read_lineup_cells(lineups_ws, row, HOME_LINEUP_COLUMNS)
            away_lineup = read_lineup_cells(lineups_ws, row, AWAY_LINEUP_COLUMNS)
            has_any_lineup = bool(home_lineup or away_lineup)

            numeric_markers = [m for m in (home_marker, away_marker) if isinstance(m, int)]
            if numeric_markers:
                current_jam_number = numeric_markers[0]
            elif home_marker == "SP" or away_marker == "SP":
                if current_jam_number == 0:
                    continue
            elif has_any_lineup:
                # Fallback for workbooks where the lineup jam marker cells are formula-driven
                # and do not have cached values. Treat the next populated lineup row as the next jam.
                current_jam_number += 1
            else:
                continue

            key = (period_number, current_jam_number)
            if key not in jams_by_key:
                continue

            jam = jams_by_key[key]

            if home_lineup:
                jam["home_lineup"] = home_lineup
            if away_lineup:
                jam["away_lineup"] = away_lineup

            if home_marker == "SP":
                jam["home_star_pass"] = True
            if away_marker == "SP":
                jam["away_star_pass"] = True


# -----------------------------
# Read penalties sheet
# -----------------------------

def build_penalty_roster_map(ws, roster_col):
    mapping = {}
    for top_row in range(PENALTY_ROSTER_START_ROW, PENALTY_ROSTER_END_ROW + 1, PENALTY_ROW_STEP):
        skater = cell_text(ws, f"{roster_col}{top_row}")
        if skater:
            mapping[top_row] = skater
    return mapping


def append_penalty_to_jam(jams_by_key, period_number, jam_number, team, skater, code):
    if jam_number is None:
        return
    key = (period_number, jam_number)
    if key not in jams_by_key:
        return

    jams_by_key[key]["penalties"].append({
        "team": team,
        "skater": skater,
        "code": code,
        "time": None,
    })


def extract_penalties_for_team(ws, jams_by_key, period_number, team, roster_col, penalty_cols):
    row_map = build_penalty_roster_map(ws, roster_col)

    for top_row, skater in row_map.items():
        bottom_row = top_row + 1
        for col in penalty_cols:
            code = cell_text(ws, f"{col}{top_row}")
            jam_number = cell_int(ws, f"{col}{bottom_row}")
            if code == "" and jam_number is None:
                continue
            if code == "":
                continue
            append_penalty_to_jam(jams_by_key, period_number, jam_number, team, skater, code)


def extract_penalties(penalties_ws, jams_by_key):
    for period_number in (1, 2):
        for team in ("home", "away"):
            extract_penalties_for_team(
                penalties_ws,
                jams_by_key,
                period_number,
                team,
                PENALTY_ROSTER_COLUMNS[period_number][team],
                PENALTY_COLUMN_GROUPS[period_number][team],
            )


# -----------------------------
# Score start/end reconstruction
# -----------------------------

def reconstruct_running_scores(jams):
    scores = {"home": 0, "away": 0}

    for jam in sorted(jams, key=lambda j: (j["period_number"], j["jam_number"])):
        jam["home_score_start"] = scores["home"]
        jam["away_score_start"] = scores["away"]

        scores["home"] += sum(jam.get("home_passes", []))
        scores["away"] += sum(jam.get("away_passes", []))

        jam["home_score_end"] = scores["home"]
        jam["away_score_end"] = scores["away"]


# -----------------------------
# Main conversion
# -----------------------------

def statsbook_to_annotation_json(xlsx_path: Path, output_path: Path | None = None):
    wb = load_workbook(xlsx_path, data_only=True)

    required_sheets = ["Score", "Lineups", "Penalties"]
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Workbook is missing required sheet: {sheet_name}")

    score_ws = wb["Score"]
    lineups_ws = wb["Lineups"]
    penalties_ws = wb["Penalties"]

    jams = extract_jams_from_score(score_ws)
    jams_by_key = index_jams(jams)

    extract_lineups(lineups_ws, jams_by_key)
    extract_penalties(penalties_ws, jams_by_key)
    reconstruct_running_scores(jams)

    result = {
        "video_file": "",
        "jams": jams,
    }

    final_output = output_json_path(xlsx_path, output_path)
    with open(final_output, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2)

    return final_output


# -----------------------------
# CLI
# -----------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Convert a filled statsbook XLSX into jam annotation JSON."
    )
    parser.add_argument("statsbook_xlsx", help="Path to the filled statsbook XLSX")
    parser.add_argument(
        "-o",
        "--output",
        help="Output JSON path. Defaults to '<xlsx filename>.jams.json'",
        default=None,
    )

    args = parser.parse_args()

    xlsx_path = Path(args.statsbook_xlsx)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Statsbook XLSX not found: {xlsx_path}")

    output_path = Path(args.output) if args.output else None
    final_output = statsbook_to_annotation_json(xlsx_path, output_path)
    print(f"Created: {final_output}")


if __name__ == "__main__":
    main()
