import json
import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# -----------------------------
# Helpers
# -----------------------------

def load_json(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def as_text(value):
    """
    Preserve skater/jammer numbers like '00', '02', '0191'.
    """
    if value is None:
        return ""
    return str(value)


def safe_list(value):
    if isinstance(value, list):
        return value
    return []


def sort_skater_numbers(numbers):
    """
    Sort skater numbers numerically where possible, while preserving
    original text formatting like leading zeros.
    """
    def sort_key(x):
        s = as_text(x).strip()
        if s.isdigit():
            return (0, int(s), s)
        return (1, s)

    cleaned = [as_text(n) for n in numbers if as_text(n).strip() != ""]
    return sorted(cleaned, key=sort_key)


def is_merged_cell(ws, coord):
    return isinstance(ws[coord], MergedCell)


def safe_write(ws, coord, value, text=False):
    """
    Write only if the target cell is a normal writable cell.
    """
    if is_merged_cell(ws, coord):
        return

    cell = ws[coord]
    cell.value = value

    if text:
        cell.number_format = "@"


# -----------------------------
# IGRF roster handling
# -----------------------------

def collect_team_skaters(jams, team_prefix):
    """
    Collect all unique skater numbers for one team from:
    - jammer field
    - lineup field
    """
    skaters = set()

    jammer_key = f"{team_prefix}_jammer"
    lineup_key = f"{team_prefix}_lineup"

    for jam in jams:
        jammer = jam.get(jammer_key)
        if jammer not in (None, ""):
            skaters.add(as_text(jammer))

        for skater in safe_list(jam.get(lineup_key)):
            if skater not in (None, ""):
                skaters.add(as_text(skater))

    return sort_skater_numbers(skaters)


def write_roster_to_igrf(ws, home_skaters, away_skaters):
    """
    IGRF tab:
    - Home: B14:B33
    - Away: I14:I33
    """
    home_start_row = 14
    away_start_row = 14
    max_rows = 20

    for i in range(max_rows):
        safe_write(ws, f"B{home_start_row + i}", "")
        safe_write(ws, f"I{away_start_row + i}", "")

    for i, skater in enumerate(home_skaters[:max_rows]):
        safe_write(ws, f"B{home_start_row + i}", skater, text=True)

    for i, skater in enumerate(away_skaters[:max_rows]):
        safe_write(ws, f"I{away_start_row + i}", skater, text=True)


# -----------------------------
# Score sheet mapping
# -----------------------------

SCORE_LAYOUT = {
    1: {"start_row": 4},
    2: {"start_row": 46},
}

HOME_COLUMNS = {
    "jam_number": "A",
    "jammer": "B",
    "lead": "D",
    "trip_cols": ["H", "I", "J", "K", "L", "M", "N", "O", "P"],
}

AWAY_COLUMNS = {
    "jam_number": "T",
    "jammer": "U",
    "lead": "W",
    "trip_cols": ["AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI"],
}


def clear_score_rows(ws, start_row, num_rows):
    """
    Clear only the rows we are actually going to use.
    """
    cols = [
        HOME_COLUMNS["jam_number"],
        HOME_COLUMNS["jammer"],
        HOME_COLUMNS["lead"],
        *HOME_COLUMNS["trip_cols"],
        AWAY_COLUMNS["jam_number"],
        AWAY_COLUMNS["jammer"],
        AWAY_COLUMNS["lead"],
        *AWAY_COLUMNS["trip_cols"],
    ]

    for row in range(start_row, start_row + num_rows):
        for col in cols:
            safe_write(ws, f"{col}{row}", "")


def write_team_score_row(ws, row, jam, team_prefix, columns):
    jammer_key = f"{team_prefix}_jammer"
    passes_key = f"{team_prefix}_passes"

    jammer = as_text(jam.get(jammer_key))
    jam_number = jam.get("jam_number", "")
    passes = safe_list(jam.get(passes_key))
    lead_jammer = jam.get("lead_jammer", "")

    safe_write(ws, f"{columns['jam_number']}{row}", jam_number)
    safe_write(ws, f"{columns['jammer']}{row}", jammer, text=True)
    safe_write(ws, f"{columns['lead']}{row}", "X" if lead_jammer == team_prefix else "")

    for trip_col in columns["trip_cols"]:
        safe_write(ws, f"{trip_col}{row}", "")

    # First pass goes in the first trip column
    for i, p in enumerate(passes):
        if i < len(columns["trip_cols"]):
            safe_write(ws, f"{columns['trip_cols'][i]}{row}", p)


def write_score_sheet(ws, jams):
    jams_by_period = {1: [], 2: []}

    for jam in jams:
        period = jam.get("period_number")
        if period in jams_by_period:
            jams_by_period[period].append(jam)

    for period in (1, 2):
        period_jams = sorted(
            jams_by_period[period],
            key=lambda j: j.get("jam_number", 0)
        )

        start_row = SCORE_LAYOUT[period]["start_row"]
        clear_score_rows(ws, start_row, len(period_jams))

        for offset, jam in enumerate(period_jams):
            row = start_row + offset
            write_team_score_row(ws, row, jam, "home", HOME_COLUMNS)
            write_team_score_row(ws, row, jam, "away", AWAY_COLUMNS)


# -----------------------------
# Lineups sheet mapping
# -----------------------------

LINEUPS_LAYOUT = {
    1: {"start_row": 4},
    2: {"start_row": 46},
}

HOME_LINEUP_COLUMNS = {
    "pivot": "G",
    "blocker_1": "K",
    "blocker_2": "O",
    "blocker_3": "S",
}

AWAY_LINEUP_COLUMNS = {
    "pivot": "AG",
    "blocker_1": "AK",
    "blocker_2": "AO",
    "blocker_3": "AS",
}


def clear_lineup_rows(ws, start_row, num_rows):
    cols = [
        *HOME_LINEUP_COLUMNS.values(),
        *AWAY_LINEUP_COLUMNS.values(),
    ]

    for row in range(start_row, start_row + num_rows):
        for col in cols:
            safe_write(ws, f"{col}{row}", "")


def write_team_lineup_row(ws, row, lineup, columns):
    # JSON lineups contain 4 skaters and the jammer is handled automatically by the sheet,
    # so write the 4 listed skaters as pivot + 3 blockers in order.
    values = [as_text(v) for v in safe_list(lineup)]

    mapped = {
        "pivot": values[0] if len(values) > 0 else "",
        "blocker_1": values[1] if len(values) > 1 else "",
        "blocker_2": values[2] if len(values) > 2 else "",
        "blocker_3": values[3] if len(values) > 3 else "",
    }

    for role, col in columns.items():
        safe_write(ws, f"{col}{row}", mapped[role], text=True)


def write_lineups_sheet(ws, jams):
    jams_by_period = {1: [], 2: []}

    for jam in jams:
        period = jam.get("period_number")
        if period in jams_by_period:
            jams_by_period[period].append(jam)

    for period in (1, 2):
        period_jams = sorted(
            jams_by_period[period],
            key=lambda j: j.get("jam_number", 0)
        )

        start_row = LINEUPS_LAYOUT[period]["start_row"]
        clear_lineup_rows(ws, start_row, len(period_jams))

        for offset, jam in enumerate(period_jams):
            row = start_row + offset
            write_team_lineup_row(ws, row, jam.get("home_lineup", []), HOME_LINEUP_COLUMNS)
            write_team_lineup_row(ws, row, jam.get("away_lineup", []), AWAY_LINEUP_COLUMNS)


# -----------------------------
# Main fill function
# -----------------------------

def fill_statsbook(json_path, template_path, output_path):
    data = load_json(json_path)
    jams = data.get("jams", [])

    wb = load_workbook(template_path)

    required_sheets = ["IGRF", "Score", "Lineups"]
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Workbook is missing required sheet: {sheet_name}")

    igrf_ws = wb["IGRF"]
    home_skaters = collect_team_skaters(jams, "home")
    away_skaters = collect_team_skaters(jams, "away")
    write_roster_to_igrf(igrf_ws, home_skaters, away_skaters)

    score_ws = wb["Score"]
    write_score_sheet(score_ws, jams)

    lineups_ws = wb["Lineups"]
    write_lineups_sheet(lineups_ws, jams)

    wb.save(output_path)


# -----------------------------
# CLI
# -----------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Fill a blank WFTDA statsbook XLSX from a jams JSON file."
    )
    parser.add_argument("json_file", help="Path to the jams JSON file")
    parser.add_argument("template_xlsx", help="Path to the blank XLSX template")
    parser.add_argument(
        "-o",
        "--output",
        help="Output XLSX path. Defaults to '<json filename>_filled.xlsx'",
        default=None,
    )

    args = parser.parse_args()

    json_path = Path(args.json_file)
    template_path = Path(args.template_xlsx)

    if not json_path.exists():
        raise FileNotFoundError(f"JSON file not found: {json_path}")

    if not template_path.exists():
        raise FileNotFoundError(f"Template XLSX not found: {template_path}")

    if args.output:
        output_path = Path(args.output)
    else:
        output_path = json_path.with_name(f"{json_path.stem}_filled.xlsx")

    fill_statsbook(json_path, template_path, output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    main()
